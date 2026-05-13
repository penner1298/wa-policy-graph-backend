import json
import csv
from duckduckgo_search import DDGS
from openpyxl import load_workbook
import time

file_path = "/Users/thejoshpenner/.openclaw/media/inbound/ba66dee1-a545-4984-b36e-8fb08c08fcc7.xlsx"
out_csv = "/Users/thejoshpenner/.openclaw/workspace/sao-scraper/new_mapped_districts.csv"

def get_entities():
    wb = load_workbook(file_path)
    sheet = wb.active
    entities = []
    
    # Assuming first row is header. Find the column with names.
    headers = [cell.value for cell in sheet[1]]
    name_idx = None
    for i, h in enumerate(headers):
        if h and 'name' in str(h).lower() or 'entity' in str(h).lower() or 'district' in str(h).lower() or 'port' in str(h).lower():
            name_idx = i
            break
            
    if name_idx is None:
        name_idx = 0 # fallback to first column
        
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[name_idx]:
            entities.append(str(row[name_idx]))
    return entities

def search_url(entity_name):
    query = f'"{entity_name}" Washington state official website'
    try:
        results = DDGS().text(query, max_results=3)
        for r in results:
            url = r['href']
            # Basic validation to avoid obvious junk
            if 'facebook' not in url and 'wikipedia' not in url and 'linkedin' not in url:
                return url
        return ""
    except Exception as e:
        print(f"Search error for {entity_name}: {e}")
        return ""

def process_entities():
    entities = get_entities()
    print(f"Found {len(entities)} entities to map.")
    
    mapped = []
    for i, entity in enumerate(entities):
        if 'wa' not in entity.lower() and 'washington' not in entity.lower():
            search_name = entity + " Washington"
        else:
            search_name = entity
            
        print(f"[{i+1}/{len(entities)}] Searching for: {search_name}")
        url = search_url(search_name)
        
        # Determine type loosely
        entity_type = "School" if "school" in entity.lower() else "Port" if "port" in entity.lower() else "District"
        
        mapped.append({
            "Name": entity,
            "Type": entity_type,
            "County": "Unknown", # Can't reliably guess county from name alone usually
            "Official_URL": url,
            "ID": entity.lower().replace(" ", "").replace(".", "").replace(",", "")
        })
        time.sleep(1.5) # Rate limit protection
        
    with open(out_csv, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=["Name", "Type", "County", "Official_URL", "ID"])
        writer.writeheader()
        writer.writerows(mapped)
        
    print(f"Finished. Saved to {out_csv}")

if __name__ == "__main__":
    process_entities()
